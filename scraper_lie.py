"""
Scraper - LIE (Lei de Incentivo ao Esporte Federal) → Supabase
Baixa o XLSX do gov.br e sobe projetos de SP no Supabase
Detecta automaticamente a URL mais recente
Autor: gerado para Felipe
Uso: python3 scraper_lie.py
"""

import os
import re
import requests
import openpyxl
from datetime import datetime
from io import BytesIO
from dotenv import load_dotenv

load_dotenv()

# ─── Config ────────────────────────────────────────────────────────────────────
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY")

PAGE_URL     = "https://www.gov.br/esporte/pt-br/acoes-e-programas/lei-de-incentivo-ao-esporte"
FALLBACK_URL = "https://www.gov.br/esporte/pt-br/acoes-e-programas/lei-de-incentivo-ao-esporte/projetos-aptos-a-captacao-atualizada-31-12-25.xlsx/@@download/file"

APENAS_SP = False

# ─── Descoberta automática da URL mais recente ─────────────────────────────────

def descobrir_url_xlsx() -> str:
    print("🔍 Buscando URL mais recente do XLSX LIE...")
    try:
        resp = requests.get(PAGE_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
        resp.raise_for_status()

        for pattern in [
            r'href=["\']([^"\']*projetos-aptos[^"\']*\.xlsx[^"\']*)["\']',
            r'href=["\']([^"\']*captacao[^"\']*\.xlsx[^"\']*)["\']',
        ]:
            matches = re.findall(pattern, resp.text, re.IGNORECASE)
            if matches:
                url = matches[0]
                if not url.startswith("http"):
                    url = "https://www.gov.br" + url
                if "@@download" not in url:
                    url = url + "/@@download/file"
                print(f"  ✅ URL encontrada: {url}")
                return url

    except requests.RequestException as e:
        print(f"  ⚠️  Erro ao buscar página: {e}")

    ano_atual = datetime.today().year
    for ano in [ano_atual, ano_atual - 1]:
        for mes in ["12", "06", "03"]:
            url_t = (
                f"https://www.gov.br/esporte/pt-br/acoes-e-programas/lei-de-incentivo-ao-esporte/"
                f"projetos-aptos-a-captacao-atualizada-{mes}-{str(ano)[2:]}.xlsx/@@download/file"
            )
            try:
                r = requests.head(url_t, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
                if r.status_code == 200:
                    print(f"  ✅ URL por variação de ano: {url_t}")
                    return url_t
            except:
                continue

    print(f"  ⚠️  Usando fallback: {FALLBACK_URL}")
    return FALLBACK_URL

# ─── Download XLSX ─────────────────────────────────────────────────────────────

def download_xlsx(url: str) -> bytes | None:
    print(f"📥 Baixando planilha...")
    try:
        resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=60)
        resp.raise_for_status()
        print(f"  ✅ Baixado: {len(resp.content)//1024}KB")
        return resp.content
    except requests.RequestException as e:
        print(f"  ❌ Erro: {e}")
        return None

# ─── Parse valor ───────────────────────────────────────────────────────────────

def parse_valor(v) -> float | None:
    """
    Converte o valor da célula para float com 2 casas decimais.
    O openpyxl já retorna float quando a célula é numérica — usa direto.
    Só faz conversão de string se necessário.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    # Se vier como string (raro), remove formatação BR
    try:
        s = str(v).strip().replace("R$", "").strip()
        # Se tem vírgula e ponto: formato BR (1.234,56)
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        # Se tem só vírgula: formato BR sem milhar (1234,56)
        elif "," in s:
            s = s.replace(",", ".")
        # Se tem só ponto: já é formato EN (1234.56) — usa direto
        return round(float(s), 2)
    except:
        return None

# ─── Parse XLSX ────────────────────────────────────────────────────────────────

def parse_xlsx(xlsx_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    print(f"  📊 Planilha: {ws.max_row} linhas × {ws.max_column} colunas")

    projetos = []
    hoje = datetime.today()

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue

        numero       = str(row[0]).strip() if row[0] else None
        processo     = str(row[1]).strip() if row[1] else None
        proponente   = str(row[2]).strip().replace("\xa0", "") if row[2] else None
        nome         = str(row[3]).strip() if row[3] else None
        manifestacao = str(row[6]).strip() if row[6] else None
        modalidade   = str(row[7]).strip() if row[7] else None
        cnpj         = str(row[8]).strip().replace("\xa0", "") if row[8] else None
        cidade       = str(row[9]).strip() if row[9] else None
        uf           = str(row[10]).strip() if row[10] else None

        if APENAS_SP and uf != "SP":
            continue

        # ✅ Usa parse_valor correto — não remove ponto de float!
        valor = parse_valor(row[11])

        # Datas — openpyxl retorna datetime direto
        data_pub = None
        if row[12]:
            if isinstance(row[12], datetime):
                data_pub = row[12].strftime("%d/%m/%Y")
            else:
                data_pub = str(row[12])[:10]

        prazo = None
        prazo_dt = None
        if row[13]:
            if isinstance(row[13], datetime):
                prazo_dt = row[13]
                prazo = row[13].strftime("%d/%m/%Y")
            else:
                prazo = str(row[13])[:10]
                for fmt in ["%d/%m/%Y", "%Y-%m-%d"]:
                    try:
                        prazo_dt = datetime.strptime(prazo, fmt)
                        break
                    except:
                        continue

        status = "em_captacao" if prazo_dt and prazo_dt >= hoje else "encerrado"

        if not numero:
            continue

        projetos.append({
            "id":              numero,
            "processo":        processo,
            "proponente":      proponente,
            "nome":            nome,
            "manifestacao":    manifestacao,
            "modalidade":      modalidade,
            "cnpj":            cnpj,
            "uf":              uf,
            "cidade":          cidade,
            "valor":           valor,
            "data_publicacao": data_pub,
            "prazo_captacao":  prazo,
            "status":          status,
        })

    return projetos

# ─── Supabase upsert ───────────────────────────────────────────────────────────

def upsert_batch(rows: list) -> bool:
    resp = requests.post(
        f"{SUPABASE_URL}/rest/v1/projetos_lie",
        headers={
            "apikey":        SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type":  "application/json",
            "Prefer":        "resolution=merge-duplicates,return=minimal",
        },
        json=rows,
        timeout=30,
    )
    if resp.status_code in (200, 201):
        return True
    print(f"  ❌ Supabase erro {resp.status_code}: {resp.text[:300]}")
    return False

# ─── Main ──────────────────────────────────────────────────────────────────────

def run():
    print("\n🚀 Iniciando scraper LIE Federal")

    url = descobrir_url_xlsx()
    xlsx_bytes = download_xlsx(url)
    if not xlsx_bytes:
        print("❌ Falha ao baixar planilha. Abortando.")
        return

    projetos = parse_xlsx(xlsx_bytes)
    filtro = "SP" if APENAS_SP else "todos os estados"
    print(f"\n📋 Projetos extraídos ({filtro}): {len(projetos)}")

    em_captacao = sum(1 for p in projetos if p["status"] == "em_captacao")
    print(f"  🟢 Em captação: {em_captacao}")
    print(f"  ⚫ Encerrados: {len(projetos) - em_captacao}")

    total_ok = 0
    for i in range(0, len(projetos), 100):
        lote = projetos[i:i+100]
        ok = upsert_batch(lote)
        if ok:
            total_ok += len(lote)
        print(f"  {'✅' if ok else '❌'} Lote {i//100 + 1}: {len(lote)} projetos")

    print(f"\n🎉 Concluído! {total_ok} projetos LIE subidos no Supabase")


if __name__ == "__main__":
    run()
